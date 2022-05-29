import openpyxl
import requests
from bs4 import BeautifulSoup
import pyautogui
from openpyxl import Workbook
from tqdm import tqdm
from datetime import date, datetime
import re
import pandas as pd

#team_list = ["KBO","한화","롯데","SSG","KIA","삼성","키움","LG","NC","두산","kt"]

team = pyautogui.prompt("구단 이름을 입력하세요")

fpath = r'C:\Users\Jeong Hyunji\Desktop\baseball\crawling\크롤링_1.xlsx'

wb = openpyxl.load_workbook(fpath)
ws = wb[team]

data = pd.read_excel(fpath, sheet_name = team)
urls = data['URL']

for index,url in tqdm(enumerate(urls)):

    i = index + 2

    if (ws.cell(row= i, column=3).value == " "):
        try:
            response = requests.get(url)
            html = response.text
            soup = BeautifulSoup(html,'html.parser')
            contents = soup.find(id= "contentDetail")
            content = str(contents)
            content = re.sub('(<([^>]+)>)', ' ', content).strip()
            content = re.sub(' +', ' ', content)
            content = re.sub('[^A-Za-z0-9가-힣]', ' ', content)
            ws.cell(row=i, column=3).value = content
            print(i, content)

        except requests.exceptions.Timeout as errd:
            print("Timeout Error : ", errd)
        except requests.exceptions.HTTPError as errb:
            print("Http Error : ", errb)
        except requests.exceptions.RequestException as erra:
            print("AnyException : ", erra)

    if (i%100==0):
        wb.save(fpath)

wb.save(fpath)