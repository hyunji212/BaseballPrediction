import openpyxl
import requests
from bs4 import BeautifulSoup
import pyautogui
from openpyxl import Workbook
from tqdm import tqdm
from datetime import date,datetime

#team_list = ["KBO","한화","롯데","SSG","KIA","삼성","키움","LG","NC","두산","kt"]
#page = {'롯데':77521}
team = pyautogui.prompt("구단 이름을 입력하세요")

fpath = r'C:\Users\Jeong Hyunji\Desktop\baseball\crawling\크롤링.xlsx'

page = 0 

today = date.today()

wb = openpyxl.load_workbook(fpath)

ws = wb[team]

if ws["A1"].value is None:
    ws["A1"] ='날짜'
    ws["B1"] ='제목'
    ws["C1"] ='내용'
    ws["D1"] ='URL'

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 120

for i in tqdm(range(1,33422, 30)):

    print(f"====={page}======입니다.")

    response = requests.get(f'https://mlbpark.donga.com/mp/b.php?p={i}&m=search&b=kbotown&query={team}&select=spf&user=')
    html = response.text
    soup = BeautifulSoup(html,'html.parser')
    titles = soup.select(".tit > .txt")
    # date 함수와 이름이 겹쳐 text_date로 사용
    text_dates = soup.select("td > span.date")    

    for title,text_date in zip(titles,text_dates):

        baseball_title = title.text
        baseball_url = title.attrs["href"]
        baseball_date = text_date.text
        if baseball_date.split("-")[0] == '2022' :
            now = datetime.strptime(baseball_date, "%Y-%m-%d").date()
            opening = datetime.strptime('2022-04-02', "%Y-%m-%d").date()

            if (now >= opening) :
            # 해당 날짜만 뽑고 십으면 if(today == now 추가)-> 현재는 0525까지 뽑음... 크롤링 도중 게시판 글 올라오는 건 막을 수 x... 그냥 drop duplicate
                ws.append([now,baseball_title,"",baseball_url])

    page += 1

wb.save(fpath)